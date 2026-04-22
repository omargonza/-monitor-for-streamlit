"""
Servicio de lectura y escritura de Excel.
"""

from datetime import datetime
from io import BytesIO
import os
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app import config


# =========================================================
# CARGA / GUARDADO BASE
# =========================================================

def cargar_excel(hoja: str, procesar_numericas: bool = False) -> pd.DataFrame:
    """Carga una hoja del Excel."""
    try:
        df = pd.read_excel(config.ARCHIVO_EXCEL, sheet_name=hoja)
    except FileNotFoundError:
        print(f"Error: No se encontró '{config.ARCHIVO_EXCEL}'")
        sys.exit(1)
    except ValueError as e:
        if "sheet" in str(e).lower():
            print(f"Error: La hoja '{hoja}' no existe")
            sys.exit(1)
        raise

    if procesar_numericas:
        columnas_numericas = [
            "costo_usd",
            "precio_reyes_ars",
            "margen_minimo_pct",
            "precio_competidor_actual_ars",
            "precio_competidor_anterior_ars",
        ]
        for col in columnas_numericas:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


def guardar_excel(df: pd.DataFrame, hoja: str) -> bool:
    """Guarda un DataFrame en una hoja del Excel. Retorna True/False."""
    try:
        with pd.ExcelWriter(
            config.ARCHIVO_EXCEL,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace",
        ) as writer:
            df.to_excel(writer, sheet_name=hoja, index=False)
        return True
    except Exception as e:
        print(f"Error al guardar: {e}")
        return False


# =========================================================
# CONFIG
# =========================================================

def cargar_config() -> dict:
    """Carga configuración desde la hoja CONFIG."""
    df = cargar_excel(config.HOJAS["CONFIG"])
    config_dict = {}

    for _, row in df.iterrows():
        clave = str(row.get("clave", "")).strip()
        valor = row.get("valor", "")
        if clave:
            config_dict[clave] = valor

    return config_dict


def actualizar_config(clave: str, valor: str) -> bool:
    """Actualiza un valor en la hoja CONFIG."""
    cfg = cargar_config()
    anterior = cfg.get(clave, "")

    valor_str = _valor_legible(valor)
    anterior_str = _valor_legible(anterior)

    # no registrar ni guardar si no cambió realmente
    if anterior_str == valor_str and clave in cfg:
        return True

    df = cargar_excel(config.HOJAS["CONFIG"])
    df["clave"] = df["clave"].astype(str).str.strip()

    mask = df["clave"] == clave
    if mask.any():
        df.loc[mask, "valor"] = valor
    else:
        new_row = pd.DataFrame({"clave": [clave], "valor": [valor]})
        df = pd.concat([df, new_row], ignore_index=True)

    resultado = guardar_excel(df, config.HOJAS["CONFIG"])

    if resultado:
        registrar_cambio_config(clave, anterior, valor)

    return resultado


def actualizar_dolar(modo: str, valor) -> bool:
    """Actualiza dólar (oficial, blue o modo) en CONFIG."""
    return actualizar_config(modo, str(valor))


# =========================================================
# PRODUCTOS
# =========================================================

def cargar_productos_monitoreo() -> pd.DataFrame:
    """Carga SOLO productos activos desde la hoja MONITOREO. Solo lectura."""
    df = cargar_excel(config.HOJAS["MONITOREO"], procesar_numericas=True)

    if "activo" not in df.columns:
        print("Error: No existe columna 'activo'")
        sys.exit(1)

    df = df[df["activo"].astype(str).str.upper().isin(["SI", "YES", "1", "TRUE"])]
    return df


def cargar_productos_completo() -> pd.DataFrame:
    """Carga TODOS los productos (activos e inactivos). Para escritura."""
    return cargar_excel(config.HOJAS["MONITOREO"], procesar_numericas=True)


def obtener_productos_todos() -> pd.DataFrame:
    """Obtiene todos los productos (activos e inactivos)."""
    return cargar_excel(config.HOJAS["MONITOREO"])


def verificar_codigo_existe(codigo: str, competidor: str | None = None) -> bool:
    """Verifica si un código ya existe en monitoreo."""
    df = cargar_productos_completo()
    codigo = str(codigo).strip().lower()

    if competidor:
        competidor = str(competidor).strip().lower()
        mask = (
            df["codigo_reyes"].astype(str).str.strip().str.lower() == codigo
        ) & (
            df["competidor"].astype(str).str.strip().str.lower() == competidor
        )
        return bool(mask.any())

    return bool(
        df["codigo_reyes"].astype(str).str.strip().str.lower().eq(codigo).any()
    )


def insertar_producto(producto: dict) -> bool:
    """Inserta un nuevo producto. Usa hoja completa."""
    df = cargar_productos_completo()
    codigo = str(producto.get("codigo_reyes", "")).strip().lower()
    competidor = str(producto.get("competidor", "")).strip().lower()

    if verificar_codigo_existe(codigo, competidor):
        return False

    new_row = pd.DataFrame([producto])
    df = pd.concat([df, new_row], ignore_index=True)

    return guardar_excel(df, config.HOJAS["MONITOREO"])


def actualizar_producto(codigo: str, competidor: str, datos: dict) -> bool:
    """Actualiza un producto existente. Usa hoja completa."""
    df = cargar_productos_completo()
    codigo_norm = str(codigo).strip().lower()
    competidor_norm = str(competidor).strip().lower()

    mask = (
        df["codigo_reyes"].astype(str).str.strip().str.lower() == codigo_norm
    ) & (
        df["competidor"].astype(str).str.strip().str.lower() == competidor_norm
    )

    if not mask.any():
        return False

    cambios = []

    for key, value in datos.items():
        if key in df.columns:
            anterior = df.loc[mask, key].values[0]
            anterior_str = _valor_legible(anterior)
            nuevo_str = _valor_legible(value)

            if anterior_str != nuevo_str:
                cambios.append((key, anterior_str, nuevo_str))

            df.loc[mask, key] = value

    resultado = guardar_excel(df, config.HOJAS["MONITOREO"])

    if resultado:
        for campo, anterior, nuevo in cambios:
            # el activo se registra aparte en cambiar_estado_producto
            if campo != "activo":
                registrar_edicion(codigo, competidor, campo, anterior, nuevo)

    return resultado


def actualizar_precio_competidor(
    codigo: str,
    competidor: str,
    nuevo_precio: float,
    mover_anterior: bool = True,
) -> bool:
    """Actualiza precio del competidor. Usa hoja completa."""
    df = cargar_productos_completo()
    codigo_norm = str(codigo).strip().lower()
    competidor_norm = str(competidor).strip().lower()

    mask = (
        df["codigo_reyes"].astype(str).str.strip().str.lower() == codigo_norm
    ) & (
        df["competidor"].astype(str).str.strip().str.lower() == competidor_norm
    )

    if not mask.any():
        return False

    anterior = df.loc[mask, "precio_competidor_actual_ars"].values[0]
    anterior_str = _valor_legible(anterior)
    nuevo_str = _valor_legible(nuevo_precio)

    # si no cambia, no guardar ni registrar
    if anterior_str == nuevo_str:
        return True

    if mover_anterior:
        df.loc[mask, "precio_competidor_anterior_ars"] = anterior

    df.loc[mask, "precio_competidor_actual_ars"] = nuevo_precio

    resultado = guardar_excel(df, config.HOJAS["MONITOREO"])

    if resultado:
        registrar_cambio_precio(codigo, competidor, anterior, nuevo_precio)

    return resultado


def cambiar_estado_producto(codigo: str, competidor: str, activo: bool) -> bool:
    """Activa o desactiva un producto. Usa hoja completa."""
    df = cargar_productos_completo()
    codigo_norm = str(codigo).strip().lower()
    competidor_norm = str(competidor).strip().lower()

    mask = (
        df["codigo_reyes"].astype(str).str.strip().str.lower() == codigo_norm
    ) & (
        df["competidor"].astype(str).str.strip().str.lower() == competidor_norm
    )

    if not mask.any():
        return False

    anterior = _valor_legible(df.loc[mask, "activo"].values[0])
    nuevo = "SI" if activo else "NO"

    if anterior == nuevo:
        return True

    resultado = actualizar_producto(codigo, competidor, {"activo": nuevo})

    if resultado:
        registrar_cambio_estado(codigo, competidor, activo)

    return resultado


# =========================================================
# HISTORIAL
# =========================================================

def _columnas_historial() -> list[str]:
    return [
        "fecha",
        "tipo_accion",
        "codigo_reyes",
        "competidor",
        "campo",
        "valor_anterior",
        "valor_nuevo",
        "usuario",
    ]


def _valor_legible(valor) -> str:
    """Normaliza valores para guardarlos/mostrarlos en historial."""
    if valor is None:
        return ""

    try:
        if pd.isna(valor):
            return ""
    except Exception:
        pass

    texto = str(valor).strip()

    if texto.lower() in {"", "nan", "none", "null"}:
        return ""

    try:
        numero = float(texto)
        if numero.is_integer():
            return str(int(numero))
        return str(numero)
    except Exception:
        return texto


def _sheet_exists(nombre_hoja: str) -> bool:
    """Verifica si existe una hoja en el Excel."""
    try:
        if not os.path.exists(config.ARCHIVO_EXCEL):
            return False

        wb = load_workbook(config.ARCHIVO_EXCEL, read_only=True)
        existe = nombre_hoja in wb.sheetnames
        wb.close()
        return existe
    except Exception:
        return False


def _inicializar_historial() -> bool:
    """
    Crea la hoja HISTORIAL si no existe.
    No debe imprimir errores por consola.
    """
    try:
        if _sheet_exists(config.HOJAS["HISTORIAL"]):
            return True

        df = pd.DataFrame(columns=_columnas_historial())

        with pd.ExcelWriter(
            config.ARCHIVO_EXCEL,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace",
        ) as writer:
            df.to_excel(writer, sheet_name=config.HOJAS["HISTORIAL"], index=False)

        return True
    except FileNotFoundError:
        return False
    except InvalidFileException:
        return False
    except Exception:
        return False


def cargar_historial(limite: int = 50) -> pd.DataFrame:
    """
    Carga el historial de cambios.
    Si la hoja no existe, la crea y devuelve DataFrame vacío.
    Nunca debe ensuciar la consola con mensajes de error.
    """
    try:
        if not _sheet_exists(config.HOJAS["HISTORIAL"]):
            _inicializar_historial()
            return pd.DataFrame(columns=_columnas_historial())

        df = pd.read_excel(config.ARCHIVO_EXCEL, sheet_name=config.HOJAS["HISTORIAL"])

        if df.empty:
            return pd.DataFrame(columns=_columnas_historial())

        if "fecha" in df.columns:
            df = df.sort_values("fecha", ascending=False)

        return df.head(limite)
    except Exception:
        return pd.DataFrame(columns=_columnas_historial())


def _registrar_evento(
    tipo_accion: str,
    codigo_reyes: str = "",
    competidor: str = "",
    campo: str = "",
    valor_anterior: str = "",
    valor_nuevo: str = "",
    usuario: str = "sistema",
) -> bool:
    """
    Registra evento en HISTORIAL.
    Si la hoja no existe, la crea antes de guardar.
    No registra eventos si no hubo cambio real.
    """
    try:
        anterior_limpio = _valor_legible(valor_anterior)
        nuevo_limpio = _valor_legible(valor_nuevo)

        # si no hubo cambio real, no registrar
        if anterior_limpio == nuevo_limpio:
            return True

        _inicializar_historial()

        if _sheet_exists(config.HOJAS["HISTORIAL"]):
            df = pd.read_excel(config.ARCHIVO_EXCEL, sheet_name=config.HOJAS["HISTORIAL"])
        else:
            df = pd.DataFrame(columns=_columnas_historial())

        # en cambios de config no tiene sentido mostrar código/competidor
        if tipo_accion == "CAMBIO_CONFIG":
            codigo_reyes = ""
            competidor = ""

        nuevo = pd.DataFrame(
            [
                {
                    "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "tipo_accion": tipo_accion,
                    "codigo_reyes": _valor_legible(codigo_reyes),
                    "competidor": _valor_legible(competidor),
                    "campo": _valor_legible(campo),
                    "valor_anterior": anterior_limpio,
                    "valor_nuevo": nuevo_limpio,
                    "usuario": _valor_legible(usuario or "sistema"),
                }
            ]
        )

        df = pd.concat([df, nuevo], ignore_index=True)

        with pd.ExcelWriter(
            config.ARCHIVO_EXCEL,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace",
        ) as writer:
            df.to_excel(writer, sheet_name=config.HOJAS["HISTORIAL"], index=False)

        return True
    except Exception:
        return False


def registrar_cambio_precio(codigo: str, competidor: str, anterior, nuevo) -> bool:
    """Registra actualización de precio competidor."""
    return _registrar_evento(
        "ACTUALIZACION_PRECIO",
        codigo_reyes=codigo,
        competidor=competidor,
        campo="precio_competidor_ars",
        valor_anterior=anterior,
        valor_nuevo=nuevo,
    )


def registrar_edicion(codigo: str, competidor: str, campo: str, anterior: str, nuevo: str) -> bool:
    """Registra edición de producto."""
    return _registrar_evento(
        "EDICION_PRODUCTO",
        codigo_reyes=codigo,
        competidor=competidor,
        campo=campo,
        valor_anterior=anterior,
        valor_nuevo=nuevo,
    )


def registrar_cambio_estado(codigo: str, competidor: str, activo: bool) -> bool:
    """Registra activación/desactivación."""
    return _registrar_evento(
        "CAMBIO_ESTADO",
        codigo_reyes=codigo,
        competidor=competidor,
        campo="activo",
        valor_anterior="NO" if activo else "SI",
        valor_nuevo="SI" if activo else "NO",
    )


def registrar_cambio_config(clave: str, anterior: str, nuevo: str) -> bool:
    """Registra cambio de configuración."""
    return _registrar_evento(
        "CAMBIO_CONFIG",
        campo=clave,
        valor_anterior=anterior,
        valor_nuevo=nuevo,
    )


# =========================================================
# IMPORTACIÓN MASIVA
# =========================================================

COLUMNAS_OBLIGATORIAS = [
    "codigo_reyes",
    "descripcion_reyes",
    "competidor",
    "costo_usd",
    "precio_reyes_ars",
    "margen_minimo_pct",
]

COLUMNAS_OPCIONALES = [
    "precio_competidor_actual_ars",
    "url_categoria_competidor",
    "referencia_manual_competidor",
    "activo",
]


def importar_productos_desde_excel(df_import: pd.DataFrame) -> dict:
    """
    Importa productos desde un DataFrame.
    
    Reglas:
    - Columnas obligatorias: codigo_reyes, descripcion_reyes, competidor, costo_usd, precio_reyes_ars, margen_minimo_pct
    - Columnas opcionales: precio_competidor_actual_ars, url_categoria_competidor, referencia_manual_competidor, activo
    - Si falta precio_competidor_actual_ars: se importa como 0 (quedará SIN_DATO)
    - Si el producto ya existe: se actualiza
    - Si el producto no existe: se crea nuevo
    - Si el mismo producto+competidor aparece dos veces: se marca como error
    
    Retorna dict con:
    - nuevos: cantidad de productos creados
    - actualizados: cantidad de productos actualizados
    - errores: cantidad de filas con error
    - omitidos: cantidad de filas omitidas
    - total: total de filas procesadas
    - sin_dato: cantidad de productos sin precio competidor
    - duplicados: cantidad de duplicados detectados en el archivo
    - detalles: lista de mensajes de resultado
    """
    resultado = {
        "nuevos": 0,
        "actualizados": 0,
        "errores": 0,
        "omitidos": 0,
        "total": 0,
        "sin_dato": 0,
        "duplicados": 0,
        "detalles": [],
    }
    
    if df_import is None or df_import.empty:
        resultado["detalles"].append("El archivo está vacío.")
        return resultado
    
    cols_present = [c.strip().lower() for c in df_import.columns]
    cols_faltantes = [c for c in COLUMNAS_OBLIGATORIAS if c.lower() not in cols_present]
    
    if cols_faltantes:
        resultado["detalles"].append(f"Columnas obligatorias faltantes: {', '.join(cols_faltantes)}")
        resultado["errores"] = len(df_import)
        return resultado
    
    df_existente = cargar_productos_completo()
    
    claves_vistas = set()
    
    for idx, row in df_import.iterrows():
        try:
            codigo = str(row.get("codigo_reyes", "")).strip()
            desc = str(row.get("descripcion_reyes", "")).strip()
            competidor = str(row.get("competidor", "")).strip()
            
            if not codigo or codigo.lower() in ["nan", "none", ""]:
                resultado["detalles"].append(f"Fila {idx+2}: Falta código.")
                resultado["errores"] += 1
                continue
                
            if not desc or desc.lower() in ["nan", "none", ""]:
                resultado["detalles"].append(f"Fila {idx+2}: Falta descripción.")
                resultado["errores"] += 1
                continue
                
            if not competidor or competidor.lower() in ["nan", "none", ""]:
                resultado["detalles"].append(f"Fila {idx+2}: Falta competidor.")
                resultado["errores"] += 1
                continue
            
            clave = (codigo.lower(), competidor.lower())
            if clave in claves_vistas:
                resultado["detalles"].append(f"Fila {idx+2}: DUPLICADO - {codigo} / {competidor} aparece más de una vez en el archivo.")
                resultado["duplicados"] += 1
                resultado["errores"] += 1
                continue
            claves_vistas.add(clave)
            
            try:
                costo_usd = float(row.get("costo_usd", 0) or 0)
            except:
                resultado["detalles"].append(f"Fila {idx+2}: Costo USD inválido.")
                resultado["errores"] += 1
                continue
                
            try:
                precio_reyes = float(row.get("precio_reyes_ars", 0) or 0)
            except:
                resultado["detalles"].append(f"Fila {idx+2}: Precio Reyes inválido.")
                resultado["errores"] += 1
                continue
                
            try:
                margen_min = float(row.get("margen_minimo_pct", 20) or 20)
            except:
                resultado["detalles"].append(f"Fila {idx+2}: Margen mínimo inválido.")
                resultado["errores"] += 1
                continue
            
            precio_comp_raw = row.get("precio_competidor_actual_ars", 0)
            try:
                precio_comp = float(precio_comp_raw) if precio_comp_raw else 0
            except:
                precio_comp = 0
            
            if precio_comp == 0 or precio_comp is None:
                resultado["sin_dato"] += 1
            
            url_cat = str(row.get("url_categoria_competidor", "") or "").strip()
            if url_cat.lower() in ["nan", "none", ""]:
                url_cat = ""
            
            ref_manual = str(row.get("referencia_manual_competidor", "") or "").strip()
            if ref_manual.lower() in ["nan", "none", ""]:
                ref_manual = ""
            
            activo_raw = row.get("activo", "SI")
            activo = "SI"
            if str(activo_raw).strip().lower() in ["no", "0", "false", "inactivo"]:
                activo = "NO"
            
            codigo_norm = codigo.lower()
            competidor_norm = competidor.lower()
            
            mask_existe = (
                df_existente["codigo_reyes"].astype(str).str.strip().str.lower() == codigo_norm
            ) & (
                df_existente["competidor"].astype(str).str.strip().str.lower() == competidor_norm
            )
            
            producto = {
                "codigo_reyes": codigo,
                "descripcion_reyes": desc,
                "competidor": competidor,
                "costo_usd": costo_usd,
                "precio_reyes_ars": precio_reyes,
                "margen_minimo_pct": margen_min,
                "precio_competidor_actual_ars": precio_comp,
                "precio_competidor_anterior_ars": 0,
                "url_categoria_competidor": url_cat,
                "referencia_manual_competidor": ref_manual,
                "activo": activo,
            }
            
            if mask_existe.any():
                idx_existente = df_existente[mask_existe].index[0]
                for col, valor in producto.items():
                    df_existente.at[idx_existente, col] = valor
                resultado["actualizados"] += 1
                resultado["detalles"].append(f"Actualizado: {codigo} / {competidor}")
            else:
                df_existente = pd.concat([df_existente, pd.DataFrame([producto])], ignore_index=True)
                resultado["nuevos"] += 1
                resultado["detalles"].append(f"Nuevo: {codigo} / {competidor}")
                
        except Exception as e:
            resultado["detalles"].append(f"Fila {idx+2}: Error - {str(e)}")
            resultado["errores"] += 1
    
    resultado["total"] = len(df_import)
    resultado["omitidos"] = resultado["errores"]
    
    if resultado["nuevos"] > 0 or resultado["actualizados"] > 0:
        guardado = guardar_excel(df_existente, config.HOJAS["MONITOREO"])
        if not guardado:
            resultado["detalles"].append("ERROR: No se pudo guardar en Excel.")
            resultado["nuevos"] = 0
            resultado["actualizados"] = 0
    
    return resultado


def generar_plantilla_importacion() -> bytes:
    """Genera un archivo Excel con la plantilla de importación."""
    data = {
        "codigo_reyes": ["EJEMPLO001", "EJEMPLO002"],
        "descripcion_reyes": ["Producto de ejemplo 1", "Producto de ejemplo 2"],
        "competidor": ["Competidor A", "Competidor B"],
        "costo_usd": [10.00, 25.50],
        "precio_reyes_ars": [15000, 38000],
        "margen_minimo_pct": [20, 20],
        "precio_competidor_actual_ars": [14500, 36000],
        "url_categoria_competidor": ["", ""],
        "referencia_manual_competidor": ["", ""],
        "activo": ["SI", "SI"],
    }
    df = pd.DataFrame(data)
    
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Productos", index=False)
    return buffer.getvalue()